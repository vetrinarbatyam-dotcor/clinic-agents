#!/usr/bin/env python3
"""Vaccine Audit — Cross-reconciliation between clinic and insurer vaccine records."""
import sys, re, os
import pandas as pd
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOLERANCE_DAYS = 7
FUZZY_THRESHOLD = 0.80
EXCLUSION = ['אולטרסאונד','ultrasound','סימפריקה','simparica','ברבקטו','bravecto','ברווקטו']
VACCINE_KEYWORDS = ['כלבת','משושה','מרובע','מתומן','תולעת','תילוע','dp plus','dp +']

def clean(val):
    if pd.isna(val): return ''
    val = str(val).replace('="','').replace('"','').replace('=','')
    val = re.sub(r'[()"]', '', val)
    return re.sub(r'\s+', ' ', val).strip()

def normalize_finals(s):
    return s.replace('\u05dd','\u05de').replace('\u05df','\u05e0').replace('\u05e3','\u05e4').replace('\u05e5','\u05e6').replace('\u05da','\u05db')

def normalize_vaccine(name):
    name = str(name).strip()
    if any(x in name for x in ['משושה','DP PLUS','DP +','פפי DP']): return 'משושה'
    if 'מתומן' in name: return 'מתומן'
    if 'מרובע' in name: return 'מרובע'
    if 'כלבת' in name: return 'כלבת'
    if 'תולעת' in name: return 'תולעת הפארק'
    if 'תילוע' in name: return 'תילוע'
    return name

def sim(a, b):
    return SequenceMatcher(None, normalize_finals(a), normalize_finals(b)).ratio()

def name_sim(a, b):
    """Try both name orders and take max."""
    direct = sim(a, b)
    rev_a = ' '.join(reversed(a.split()))
    rev_b = ' '.join(reversed(b.split()))
    return max(direct, sim(rev_a, b), sim(a, rev_b))

def strip_english(s):
    return re.sub(r'[a-zA-Z]', '', s).strip()

def vaccines_compatible(v1, v2):
    if v1 == v2: return True
    if {v1, v2} == {'מתומן', 'משושה'}: return True
    return False

def is_excluded(name):
    nl = name.lower()
    return any(ex.lower() in nl for ex in EXCLUSION)

# ============================================================
# FILE LOADING — auto-detect structure
# ============================================================
def load_clinic(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        for enc in ['utf-16','utf-8-sig','utf-8','cp1255']:
            for sep in ['\t',',']:
                try:
                    df = pd.read_csv(path, encoding=enc, sep=sep, nrows=3)
                    if len(df.columns) > 2:
                        df = pd.read_csv(path, encoding=enc, sep=sep)
                        break
                except: continue
            else: continue
            break
    else:
        df = pd.read_excel(path)

    col_map = {}
    for c in df.columns:
        cl = str(c).strip()
        if 'שם הלקוח' in cl or cl == 'לקוח': col_map['client'] = c
        elif 'שם החיה' in cl or cl == 'חיה' or cl == 'שם בעח': col_map['animal'] = c
        elif 'תאריך' in cl and 'date' not in col_map: col_map['date'] = c
        elif 'חיסון' in cl or 'פריט' in cl: col_map['vaccine'] = c

    df['client'] = df[col_map['client']].apply(clean)
    df['animal'] = df[col_map.get('animal', col_map['client'])].apply(clean)
    df['date'] = pd.to_datetime(df[col_map['date']].apply(clean), format='%d/%m/%Y', errors='coerce')
    if df['date'].isna().all():
        df['date'] = pd.to_datetime(df[col_map['date']].apply(clean), format='%d-%m-%Y', errors='coerce')
    if df['date'].isna().all():
        df['date'] = pd.to_datetime(df[col_map['date']], errors='coerce')
    df['vaccine'] = df[col_map['vaccine']].apply(clean)
    df = df.dropna(subset=['date'])
    df = df[df['vaccine'].str.strip() != '']
    if df.empty:
        return pd.DataFrame(columns=['client','animal','date','vaccine','vaccine_norm'])
    df = df[~df['vaccine'].apply(is_excluded)].copy()
    if df.empty:
        return pd.DataFrame(columns=['client','animal','date','vaccine','vaccine_norm'])
    df['vaccine_norm'] = df['vaccine'].apply(normalize_vaccine)
    return df[['client','animal','date','vaccine','vaccine_norm']].reset_index(drop=True)

def load_insurer(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        for enc in ['utf-16','utf-8-sig','utf-8','cp1255']:
            for sep in ['\t',',']:
                try:
                    df = pd.read_csv(path, encoding=enc, sep=sep, nrows=3)
                    if len(df.columns) > 2:
                        df = pd.read_csv(path, encoding=enc, sep=sep)
                        break
                except: continue
            else: continue
            break
    else:
        df = pd.read_excel(path)

    has_split_name = any('שם פרטי' in str(c) for c in df.columns)
    col_map = {}
    for c in df.columns:
        cl = str(c).strip()
        if 'שם פרטי' in cl: col_map['first'] = c
        elif 'שם משפחה' in cl or 'משפחה' in cl: col_map['last'] = c
        elif 'שם הלקוח' in cl or 'לקוח' in cl or cl == 'שם לקוח': col_map['client'] = c
        elif 'שם החיה' in cl or cl == 'שם בעח' or cl == 'חיה': col_map['animal'] = c
        elif 'תאריך' in cl: col_map['date'] = c
        elif 'רשימת טיפולים' in cl or 'טיפול' in cl or 'פריט' in cl: col_map['treatments'] = c

    if has_split_name:
        df['client'] = (df[col_map['first']].astype(str).str.strip() + ' ' + df[col_map['last']].astype(str).str.strip()).apply(clean)
    else:
        df['client'] = df[col_map['client']].apply(clean)

    has_animal = 'animal' in col_map
    df['date'] = pd.to_datetime(df[col_map['date']], errors='coerce')
    if df['date'].isna().all():
        df['date'] = pd.to_datetime(df[col_map['date']].apply(clean), format='%d-%m-%Y', errors='coerce')

    rows = []
    for _, row in df.iterrows():
        raw = str(row[col_map['treatments']])
        if raw in ['nan','NaT','']: continue
        for t in raw.split(','):
            t = clean(t)
            if not t: continue
            if any(kw in t for kw in VACCINE_KEYWORDS):
                if is_excluded(t): continue
                animal = strip_english(clean(str(row[col_map['animal']]))) if has_animal else ''
                rows.append({'client': row['client'], 'animal': animal,
                             'date': row['date'], 'vaccine': t,
                             'vaccine_norm': normalize_vaccine(t)})
    if not rows:
        return pd.DataFrame(columns=['client','animal','date','vaccine','vaccine_norm'])
    return pd.DataFrame(rows).reset_index(drop=True)

# ============================================================
# MATCHING ENGINE
# ============================================================
def run_matching(clinic, insurer):
    if clinic.empty and insurer.empty:
        empty_t1 = pd.DataFrame(columns=['שם לקוח','שם חיה','תאריך','חיסון (מנורמל)','חיסון (מקורי)','הערה'])
        clinic['matched'] = pd.Series(dtype=bool)
        clinic['match_level'] = pd.Series(dtype=str)
        clinic['match_note'] = pd.Series(dtype=str)
        return empty_t1, pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), clinic
    clinic['matched'] = False
    clinic['match_level'] = ''
    clinic['match_note'] = ''
    insurer['matched'] = False

    has_animal = 'animal' in insurer.columns and insurer['animal'].str.strip().ne('').any()
    table2, table3 = [], []

    # Level 1: Exact name + vaccine + date
    for ci in clinic.index:
        if clinic.at[ci,'matched']: continue
        for ii in insurer.index:
            if insurer.at[ii,'matched']: continue
            if not vaccines_compatible(clinic.at[ci,'vaccine_norm'], insurer.at[ii,'vaccine_norm']): continue
            if abs((clinic.at[ci,'date'] - insurer.at[ii,'date']).days) > TOLERANCE_DAYS: continue
            ns = name_sim(clinic.at[ci,'client'], insurer.at[ii,'client'])
            if ns < FUZZY_THRESHOLD: continue
            # Animal check if available
            if has_animal and insurer.at[ii,'animal']:
                asim = sim(strip_english(clinic.at[ci,'animal']), insurer.at[ii,'animal'])
                if asim < 0.50: continue
            clinic.at[ci,'matched'] = True
            clinic.at[ci,'match_level'] = 'L1-Exact' if ns >= 0.98 else 'L2-Fuzzy'
            insurer.at[ii,'matched'] = True
            if ns < 0.98:
                table3.append({
                    'לקוח_מרפאה': clinic.at[ci,'client'], 'חיה_מרפאה': clinic.at[ci,'animal'],
                    'לקוח_מרפאט': insurer.at[ii,'client'],
                    'תאריך_מרפאה': clinic.at[ci,'date'].strftime('%d/%m/%Y'),
                    'תאריך_מרפאט': insurer.at[ii,'date'].strftime('%d/%m/%Y'),
                    'חיסון': clinic.at[ci,'vaccine_norm']})
            break

    # Level 3: Owner Switch — same vaccine + date ±1 + different name
    for ci in clinic.index:
        if clinic.at[ci,'matched']: continue
        for ii in insurer.index:
            if insurer.at[ii,'matched']: continue
            if not vaccines_compatible(clinic.at[ci,'vaccine_norm'], insurer.at[ii,'vaccine_norm']): continue
            if abs((clinic.at[ci,'date'] - insurer.at[ii,'date']).days) > 1: continue
            ns = name_sim(clinic.at[ci,'client'], insurer.at[ii,'client'])
            if ns >= FUZZY_THRESHOLD: continue  # already handled in L1/L2
            # Require animal similarity for owner switch
            if has_animal and insurer.at[ii,'animal']:
                asim = sim(strip_english(clinic.at[ci,'animal']), insurer.at[ii,'animal'])
                if asim < 0.70: continue
            elif not has_animal:
                continue  # without animal name, owner switch is unreliable
            clinic.at[ci,'matched'] = True
            clinic.at[ci,'match_level'] = 'L3-OwnerSwitch'
            clinic.at[ci,'match_note'] = insurer.at[ii,'client']
            insurer.at[ii,'matched'] = True
            table2.append({
                'לקוח_מרפאה': clinic.at[ci,'client'], 'לקוח_מרפאט': insurer.at[ii,'client'],
                'חיה': clinic.at[ci,'animal'], 'תאריך': clinic.at[ci,'date'].strftime('%d/%m/%Y'),
                'חיסון': clinic.at[ci,'vaccine_norm'], 'הערה': 'זוהה לפי חיסון+תאריך+חיה'})
            break

    # Reverse orphans
    table4 = []
    for ii in insurer.index:
        if insurer.at[ii,'matched']: continue
        table4.append({
            'שם לקוח (מרפאט)': insurer.at[ii,'client'],
            'תאריך': insurer.at[ii,'date'].strftime('%d/%m/%Y'),
            'חיסון': insurer.at[ii,'vaccine_norm'],
            'חיסון (מקורי)': insurer.at[ii,'vaccine'],
            'הערה': 'קיים במרפאט ללא התאמה במרפאה'})

    unmatched = clinic[~clinic['matched']][['client','animal','date','vaccine_norm','vaccine']].copy()
    unmatched.columns = ['שם לקוח','שם חיה','תאריך','חיסון (מנורמל)','חיסון (מקורי)']
    if not unmatched.empty and pd.api.types.is_datetime64_any_dtype(unmatched['תאריך']):
        unmatched['תאריך'] = unmatched['תאריך'].dt.strftime('%d/%m/%Y')
    unmatched['הערה'] = 'לא נמצא בביטוח'

    return unmatched, pd.DataFrame(table2), pd.DataFrame(table3), pd.DataFrame(table4), clinic

# ============================================================
# EXCEL OUTPUT
# ============================================================
def write_excel(t1, t2, t3, t4, clinic, output_path):
    wb = Workbook()
    hf = Font(bold=True, size=11, color='FFFFFF', name='Arial')
    df_ = Font(size=10, name='Arial')
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    ca = Alignment(horizontal='center', vertical='center')
    ra = Alignment(horizontal='right', vertical='center')

    def sheet(ws, title, tc, hc, headers, data, rf):
        ws.sheet_view.rightToLeft = True
        nc = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
        ws['A1'].value = title; ws['A1'].font = Font(bold=True, size=14, color=tc, name='Arial'); ws['A1'].alignment = ca
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=i, value=h)
            c.font = hf; c.fill = PatternFill('solid', fgColor=hc); c.alignment = ca; c.border = bd
        for ri, row in enumerate(data, 3):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = df_; c.alignment = ra; c.border = bd; c.fill = PatternFill('solid', fgColor=rf)
        for i in range(1, nc + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

    ws1 = wb.active; ws1.title = 'חיסונים חסרים'
    h1 = ['שם לקוח','שם חיה','תאריך','חיסון (מנורמל)','חיסון (מקורי)','הערה']
    d1 = [list(r) for _, r in t1.iterrows()]
    sheet(ws1, f'חיסונים חסרים — {len(t1)} רשומות', 'E36C09', 'E36C09', h1, d1, 'FCE4D6')

    ws2 = wb.create_sheet('חריגות בעלים')
    h2 = ['לקוח (מרפאה)','לקוח (מרפאט)','שם חיה','תאריך','חיסון','הערה']
    d2 = [list(r) for _, r in t2.iterrows()] if len(t2) > 0 else []
    sheet(ws2, f'חריגות בעלים — {len(t2)} רשומות', 'BF8F00', 'BF8F00', h2, d2, 'FFF2CC')

    ws3 = wb.create_sheet('דוח דלתא')
    h3 = ['לקוח (מרפאה)','חיה (מרפאה)','לקוח (מרפאט)','תאריך מרפאה','תאריך מרפאט','חיסון']
    d3 = [list(r) for _, r in t3.iterrows()] if len(t3) > 0 else []
    sheet(ws3, f'דוח דלתא — {len(t3)} רשומות', '006100', '006100', h3, d3, 'C6EFCE')

    ws4 = wb.create_sheet('יתומים במרפאט')
    h4 = ['שם לקוח (מרפאט)','תאריך','חיסון','חיסון (מקורי)','הערה']
    d4 = [list(r) for _, r in t4.iterrows()] if len(t4) > 0 else []
    sheet(ws4, f'יתומים במרפאט — {len(t4)} רשומות', '7030A0', '7030A0', h4, d4, 'E8D5F5')

    # Summary
    ws5 = wb.create_sheet('סיכום'); ws5.sheet_view.rightToLeft = True
    ws5.column_dimensions['A'].width = 45; ws5.column_dimensions['B'].width = 15
    ws5.merge_cells('A1:B1')
    ws5['A1'].value = 'סיכום ביקורת חיסונים'; ws5['A1'].font = Font(bold=True, size=14, name='Arial'); ws5['A1'].alignment = ca
    total = len(clinic)
    l1 = len(clinic[clinic['match_level']=='L1-Exact'])
    l2 = len(clinic[clinic['match_level']=='L2-Fuzzy'])
    l3 = len(clinic[clinic['match_level']=='L3-OwnerSwitch'])
    summary = [
        ('סה"כ חיסונים במרפאה (אחרי סינון)', total),
        ('התאמה מלאה (L1)', l1), ('התאמה פונטית (L2)', l2),
        ('חריגת בעלים (L3)', l3), ('סה"כ הותאמו', l1+l2+l3),
        ('', ''), ('חסרים במרפאט', len(t1)), ('יתומים במרפאט', len(t4)),
    ]
    for i, (label, val) in enumerate(summary, 3):
        c1 = ws5.cell(row=i, column=1, value=label)
        c2 = ws5.cell(row=i, column=2, value=val)
        bold = 'סה"כ' in str(label) or 'חסרים' in str(label) or 'יתומים' in str(label)
        c1.font = Font(size=11, name='Arial', bold=bold)
        c2.font = Font(size=11, name='Arial', bold=True); c2.alignment = ca
        if 'חסרים' in str(label):
            c1.font = Font(size=11, name='Arial', bold=True, color='CC0000')
            c2.font = Font(size=11, name='Arial', bold=True, color='CC0000')

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Total: {total} | L1: {l1} | L2: {l2} | L3: {l3} | Missing: {len(t1)} | Orphans: {len(t4)}")

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Usage: vaccine_audit.py <clinic_csv> <insurer_csv_or_xlsx> <output.xlsx>")
        sys.exit(1)
    clinic = load_clinic(sys.argv[1])
    insurer = load_insurer(sys.argv[2])
    t1, t2, t3, t4, clinic = run_matching(clinic, insurer)
    write_excel(t1, t2, t3, t4, clinic, sys.argv[3])
