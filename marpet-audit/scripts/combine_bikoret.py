"""Combine vaccine and treatment audit XLSXs into bikoret_{date}.xlsx with 6 sheets.

Usage: python combine_bikoret.py YYYY-MM-DD
Output: bikoret_{date}.xlsx (falls back to bikoret_{date}_new.xlsx if file is locked)
Prints: SUMMARY|lost=N|lost_val=X|missing_vacc=Y|orphans=Z
"""
import sys
import os
import openpyxl
from openpyxl.styles import Font, Alignment
from copy import copy

if len(sys.argv) >= 4:
    # New invocation: combine_bikoret.py <vaccine_audit.xlsx> <treatment_audit.xlsx> <output.xlsx>
    VACC_FILE  = sys.argv[1]
    TREAT_FILE = sys.argv[2]
    OUT_FILE   = sys.argv[3]
    OUT_TMP    = OUT_FILE.replace('.xlsx', '_new.xlsx')
    # Extract date from output filename (bikoret_YYYY-MM-DD.xlsx)
    import re
    m = re.search(r'(\d{4}-\d{2}-\d{2})', os.path.basename(OUT_FILE))
    DATE = m.group(1) if m else "unknown"
else:
    DATE = sys.argv[1] if len(sys.argv) > 1 else "2026-03-20"
    WORK_DIR = os.environ.get("MARPAT_WORK_DIR", "/home/claude-user/marpat")
    VACC_FILE  = f"{WORK_DIR}/vaccine_audit_{DATE}.xlsx"
    TREAT_FILE = f"{WORK_DIR}/treatment_audit_{DATE}.xlsx"
    OUT_FILE   = f"{WORK_DIR}/bikoret_{DATE}.xlsx"
    OUT_TMP    = f"{WORK_DIR}/bikoret_{DATE}_new.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────

def copy_sheet(src_ws, dst_wb, sheet_name, tab_color):
    dst_ws = dst_wb.create_sheet(title=sheet_name)
    dst_ws.sheet_properties.tabColor = tab_color
    for row in src_ws.iter_rows():
        for cell in row:
            nc = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font          = copy(cell.font)
                nc.fill          = copy(cell.fill)
                nc.border        = copy(cell.border)
                nc.alignment     = copy(cell.alignment)
                nc.number_format = cell.number_format
    for col in src_ws.column_dimensions:
        dst_ws.column_dimensions[col].width = src_ws.column_dimensions[col].width
    for row in src_ws.row_dimensions:
        dst_ws.row_dimensions[row].height = src_ws.row_dimensions[row].height
    for mc in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(mc))
    return dst_ws


def append_rows(src_ws, dst_ws, skip_header=True):
    """Append all rows from src_ws to dst_ws (after its last row)."""
    start_row = dst_ws.max_row + 1
    first = True
    for row in src_ws.iter_rows():
        if first and skip_header:
            first = False
            continue
        for cell in row:
            nc = dst_ws.cell(row=start_row + cell.row - (2 if skip_header else 1),
                              column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font      = copy(cell.font)
                nc.fill      = copy(cell.fill)
                nc.alignment = copy(cell.alignment)


def get_stats(ws):
    """Return {label: value} from a 2-column summary sheet."""
    stats = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] and len(row) > 1 and row[1] is not None:
            stats[str(row[0]).strip()] = row[1]
    return stats


# ── load sources ─────────────────────────────────────────────────────────────

wb_v = openpyxl.load_workbook(VACC_FILE)
wb_t = openpyxl.load_workbook(TREAT_FILE)
wb_out = openpyxl.Workbook()
wb_out.remove(wb_out.active)

# 1. תביעות אבודות — treatment audit
copy_sheet(wb_t["תביעות אבודות"], wb_out, "תביעות אבודות", "C00000")

# 2. חיסונים חסרים — vaccine audit
copy_sheet(wb_v["חיסונים חסרים"], wb_out, "חיסונים חסרים", "ED7D31")

# 3. יתומים במרפאט — combined orphans (vaccine first, treatment appended)
ws_comb = copy_sheet(wb_v["יתומים במרפאט"], wb_out, "יתומים במרפאט", "7030A0")
append_rows(wb_t["יתומים במרפאט"], ws_comb, skip_header=True)

# 4. חשד להתאמה — treatment audit
copy_sheet(wb_t["חשד להתאמה"], wb_out, "חשד להתאמה", "FFC000")

# 5. דוח דלתא — combined (vaccine first, treatment appended)
ws_delta = copy_sheet(wb_v["דוח דלתא"], wb_out, "דוח דלתא", "00B050")
append_rows(wb_t["דוח דלתא"], ws_delta, skip_header=True)

# 6. סיכום — combined stats
v_stats = get_stats(wb_v["סיכום"])
t_stats = get_stats(wb_t["סיכום"])

ws_sum = wb_out.create_sheet(title="סיכום")
ws_sum.sheet_properties.tabColor = "FFFFFF"

# Format date for display: YYYY-MM-DD → DD/MM/YYYY
display_date = "-".join(reversed(DATE.split("-")))
ws_sum["A1"] = f"סיכום ביקורת יומית — {display_date}"
ws_sum["A1"].font      = Font(bold=True, size=14)
ws_sum["A1"].alignment = Alignment(horizontal="right")

row = 3
ws_sum.cell(row=row, column=1, value="📋 טיפולים").font = Font(bold=True, size=12)
row += 1
for k, v in t_stats.items():
    ws_sum.cell(row=row, column=1, value=k)
    ws_sum.cell(row=row, column=2, value=v)
    row += 1

row += 1
ws_sum.cell(row=row, column=1, value="💉 חיסונים").font = Font(bold=True, size=12)
row += 1
for k, v in v_stats.items():
    ws_sum.cell(row=row, column=1, value=k)
    ws_sum.cell(row=row, column=2, value=v)
    row += 1

ws_sum.column_dimensions["A"].width = 38
ws_sum.column_dimensions["B"].width = 20

# ── save (handle locked file) ─────────────────────────────────────────────────

saved_to = OUT_FILE
try:
    wb_out.save(OUT_FILE)
except PermissionError:
    wb_out.save(OUT_TMP)
    saved_to = OUT_TMP
    print(f"WARNING: {OUT_FILE} is locked (open in Excel?). Saved to {OUT_TMP}")

print(f"Saved: {saved_to}")

# ── extract key metrics for WhatsApp notification ─────────────────────────────
# Exact Hebrew keys from audit scripts:
# Treatment: "תביעות אבודות (חסרים בביטוח)", "שווי תביעות אבודות (₪)", "יתומים במרפאט"
# Vaccine:   "חסרים במרפאט", "יתומים במרפאט"

lost       = t_stats.get("תביעות אבודות (חסרים בביטוח)", 0)
lost_val   = t_stats.get("שווי תביעות אבודות (₪)", 0)
miss_vacc  = v_stats.get("חסרים במרפאט", 0)
orp_t      = int(t_stats.get("יתומים במרפאט", 0) or 0)
orp_v      = int(v_stats.get("יתומים במרפאט", 0) or 0)
orphans    = orp_t + orp_v

print(f"SUMMARY|lost={lost}|lost_val={lost_val}|missing_vacc={miss_vacc}|orphans={orphans}")
print(f"FILE|{saved_to}")
