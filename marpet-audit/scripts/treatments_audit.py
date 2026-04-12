#!/usr/bin/env python3
"""Treatments Audit — Find lost claims between clinic visits and insurer (מרפאט)."""
import sys, re, os
import pandas as pd
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TOLERANCE_DAYS = 7
REVERSE_TOLERANCE = 2
FUZZY = 0.80

# Generic insurer items that could match various specific clinic items
# These are matched LAST to prefer specific matches
GENERIC_NORMS = {'זריקה_כללית', 'בדיקה', 'ביקור', 'בדיקות דם'}

# Visit injections — interchangeable 1:1 within same visit
# (clinic and insurer may record different drug names for the same billed event)
VISIT_INJECTIONS = {
    'זריקת אנטיביוטיקה', 'זריקת כאבים', 'זריקת סטרואידים',
    'זריקת דיפנהידרמין', 'זריקת פמוטידין', 'זריקת פרמין',
}

# ============================================================
# EXCLUSION LIST — non-insured items (meds, food, supplies)
# ============================================================
EXCLUSION = [
    'סימפריקה','חבילה סימפריקה','ברווקטו','ברבקטו',
    'אמוקסיקלאב','סינולוקס','מטרונידזול','מטרו ','גבאפנטין','אפוקוול',
    'דרמיפרד','טרזודיל','אנרופלוקססין','סטומורג','קלינדמיצין','דנמרין',
    'אנלדקס','אומפרדקס','פוסיד','פרוויקוקס','טרים p','ונקטיל','דיפירון',
    'ספירולונקטון','קרדיושור','סרניה 24','בולוביה','אוריזון',
    'gi biome','ha כלב','id low fat','id biome','גסטרואינטסטינל','פאוץ',
    'פחית שימורים','היפו','שמפו','דבש מנוקה','משחת עור','משחת עיניים',
    'ocil','מיטקס','איזוטיק','פחם פעיל','מנוי בוגרים',
    'מריטת פרווה','סרום ',
    # Non-insured injections (long-acting / biological / immunology)
    'סולנסיה','ציטופוינט','cytopoint','ליברלה','librela','קונבניה','convenia',
    # Referrals — just referral letters, not actual treatments
    'הפניה',
]

def clean(val):
    if pd.isna(val): return ''
    val = str(val).replace('="','').replace('"','').replace('=','')
    val = re.sub(r'[()]', '', val)
    return re.sub(r'\s+', ' ', val).strip()

def normalize_finals(s):
    """Normalize Hebrew final letters for comparison."""
    return s.replace('ם','מ').replace('ן','נ').replace('ף','פ').replace('ץ','צ').replace('ך','כ')

def sim(a, b):
    return SequenceMatcher(None, normalize_finals(a), normalize_finals(b)).ratio()

def is_excluded(name):
    n = name.lower()
    return any(ex.lower() in n for ex in EXCLUSION)

ANIMAL_MIN_SIM = 0.55  # min similarity between animal names to allow a match

def animal_compatible(c_animal, i_animal):
    """Return True if animal names are compatible.
    If either side is missing → don't block the match.
    If both present → require at least ANIMAL_MIN_SIM similarity to avoid cross-pet matches."""
    if not c_animal or not i_animal:
        return True
    return sim(c_animal, i_animal) >= ANIMAL_MIN_SIM

# ============================================================
# TREATMENT NORMALIZATION
# ============================================================
def normalize_treatment(name):
    n = name.strip()
    nl = n.lower()

    # *** Injections — checked FIRST, before visit, to handle "זריקה במהלך ביקור X" ***
    # --- Visit injections (interchangeable, same billing code) ---
    if 'אנטיביוטיקה' in nl and ('זריק' in nl or 'injection' in nl):
        return 'זריקת אנטיביוטיקה'
    if 'סטרואידים' in nl and 'זריק' in nl:
        return 'זריקת סטרואידים'
    if ('כאבים' in nl or 'כאב' in nl) and 'זריק' in nl:
        return 'זריקת כאבים'
    if 'זריקה נגד כאבים' in nl:
        return 'זריקת כאבים'
    if ('פמוטידין' in nl or 'פמויטידן' in nl) and 'זריק' in nl:  # spelling variants
        return 'זריקת פמוטידין'
    if ('דיפנהידרמין' in nl or 'דיפנדרמין' in nl) and 'זריק' in nl:
        return 'זריקת דיפנהידרמין'
    if ('פרמין' in nl or 'פראמין' in nl) and 'זריק' in nl:
        return 'זריקת פרמין'
    # --- Unique injections (specific price, NOT interchangeable) ---
    if 'סרניה' in nl and 'זריק' in nl:
        return 'זריקת סרניה'
    if 'בוטורפנול' in nl:
        return 'זריקת בוטורפנול'
    if ('דפומדרול' in nl or 'דפומדרל' in nl or 'דקסדומיטור' in nl):
        return 'זריקת דפומדרול'
    if 'אפומורפין' in nl:
        return 'זריקת אפומורפין'
    if nl in ['זריקות','זריקה']:
        return 'זריקה_כללית'

    # Visit — after injection-specific checks
    visit_words = ['בדיקה רפואית','ביקורת','ביקור','מעקב','חירום']
    no_visit = ['אולטרסאונד','שתן','דם','קרדיו','אולטראסאונד','לב','שיניים','פלורסין',
                'אוזניים','שקים','אבצס','הקאה']
    if any(v in nl for v in visit_words) and not any(x in nl for x in no_visit):
        return 'ביקור'

    # Blood bundles
    if any(b in nl for b in ['ספירת דם+כימיה','ספירת דם + פאנל','בדיקת דם מלאה','בדיקות דם']):
        return 'BUNDLE:FULL'
    if any(b in nl for b in ['כימיה מורחבת','elemnt 16','כימיה 21']):
        return 'BUNDLE:EXTENDED'
    if any(b in nl for b in ['chem 10','כימיה 10','לפני הרדמה','pre-anesthetic']):
        return 'BUNDLE:PRE'
    if 'cbc' in nl and 'כימיה 17' in nl:
        return 'BUNDLE:FULL'
    if 'cbc' in nl and 'משטח' in nl:
        return 'ספירה+משטח'
    if 'כימיה 17' in nl or 'יתרה כימיה' in nl:
        return 'כימיה 17'
    if nl == 'crp' or nl == 'c.r.p':
        return 'CRP'
    if any(x in nl for x in ['cpli','cpl כמותי','פנל תפקודי לבלב','פאנל לבלב']):
        return 'פאנל לבלב'

    # Procedures
    if 'אולטרסאונד' in nl or 'אולטראסאונד' in nl:
        return 'אולטרסאונד'
    if 'ריקון בלוטות' in nl or 'ניקוז שקים' in nl or nl == 'שקים':
        return 'ניקוז שקים'
    if 'שטיפת אוזניים' in nl:
        return 'שטיפת אוזניים'
    if 'חבישה' in nl:
        return 'חבישה'
    if 'צילום' in nl:
        return 'צילום'
    if 'עירוי' in nl:
        return 'עירוי'
    if 'בדיקת שתן' in nl or 'we cysto' in nl:
        return 'בדיקת שתן'
    if 'צביעת פלורסין' in nl:
        return 'צביעת פלורסין'
    if 'הרגעה' in nl or 'טשטוש' in nl:
        return 'הרגעה'
    if 'ניקוז אבצס' in nl:
        return 'ניקוז אבצס'
    if 'ניקוי שיניים' in nl:
        return 'ניקוי שיניים'
    if 'סירוס' in nl or 'עיקור' in nl:
        return 'ניתוח'
    if 'השריית הקאה' in nl:
        return 'השריית הקאה'
    if any(x in nl for x in ['קרדיולוג','אקו לב','קרדיולוגי','bnp']):
        return 'מומחה לב'
    if any(x in nl for x in ['אונקולוג']):
        return 'מומחה אונקולוג'
    if any(x in nl for x in ['נוירולוג']):
        return 'מומחה נוירולוג'
    if 'מומחה' in nl:
        return 'מומחה'
    if any(x in nl for x in ['fna','משטח','ציטולוגיה']):
        return 'משטחים'
    if 'פרוקטוזמין' in nl or 'אמוניה' in nl or 'sdma' in nl:
        return 'בדיקות דם מיוחדות'
    if 'דם' in nl and nl not in ['כימיה 17']:
        return 'בדיקות דם'
    return n

# ============================================================
# BUNDLE LOGIC
# ============================================================
BUNDLE_COMPONENTS = {
    'BUNDLE:FULL': {'ספירה', 'כימיה 17'},
    'BUNDLE:EXTENDED': {'ספירה', 'כימיה 17', 'אלקטרוליטים'},
    'BUNDLE:PRE': {'כימיה 10', 'לפני הרדמה'},
}

# What normalized names count as a bundle component
COMPONENT_TO_BUNDLE_PART = {
    'כימיה 17': 'כימיה 17',
    'ספירה+משטח': 'ספירה',
    'CRP': None,  # standalone
}

def is_bundle(norm):
    return norm.startswith('BUNDLE:')

def bundle_contains(bundle_norm, component_norm):
    """Check if a component is part of a bundle."""
    parts = BUNDLE_COMPONENTS.get(bundle_norm, set())
    # Direct membership
    if component_norm in parts:
        return True
    # Partial matches
    cnl = component_norm.lower()
    for p in parts:
        if p in cnl or cnl in p:
            return True
    return False

def treatments_match(c_norm, i_norm):
    if c_norm == i_norm:
        return True
    # Generic injection matches ONE specific injection (1:1 only)
    if i_norm == 'זריקה_כללית' and 'זריקת' in c_norm:
        return True
    if c_norm == 'זריקה_כללית' and 'זריקת' in i_norm:
        return True
    # Bare ingredient = injection of that ingredient (1:1)
    bare_to_injection = {
        'פמוטידין': 'זריקת פמוטידין',
        'סרניה': 'זריקת סרניה',
        'כאבים': 'זריקת כאבים',
        'אנטיביוטיקה': 'זריקת אנטיביוטיקה',
    }
    if i_norm in bare_to_injection and c_norm == bare_to_injection[i_norm]:
        return True
    if c_norm in bare_to_injection and i_norm == bare_to_injection[c_norm]:
        return True
    # Visit injections are interchangeable 1:1 (same billing event, different drug names)
    if c_norm in VISIT_INJECTIONS and i_norm in VISIT_INJECTIONS:
        return True
    # Bundle equivalences
    if {c_norm, i_norm} & {'בדיקות דם'} and {c_norm, i_norm} & {'BUNDLE:FULL', 'BUNDLE:EXTENDED'}:
        return True
    if {c_norm, i_norm} == {'BUNDLE:FULL', 'BUNDLE:EXTENDED'}:
        return True
    # "בדיקה" alone can match "ביקור"
    if {c_norm, i_norm} == {'בדיקה', 'ביקור'}:
        return True
    # Fuzzy on normalized names
    if sim(c_norm, i_norm) >= 0.85:
        return True
    return False

# ============================================================
# INSURER EXPANSION — expand "זריקות (X, Y)" into components
# ============================================================
def expand_insurer_treatment(raw):
    """Expand combined insurer treatments into individual items."""
    n = raw.strip()
    # Pattern: "זריקות (X, Y)"  or "זריקות (X, Y, Z)"
    m = re.match(r'זריקות?\s*\(([^)]+)\)', n)
    if m:
        parts = [p.strip() for p in m.group(1).split(',')]
        expanded = []
        for p in parts:
            pl = p.strip().lower()
            if 'כאב' in pl: expanded.append('זריקת כאבים')
            elif 'אנטיביוטיקה' in pl: expanded.append('זריקת אנטיביוטיקה')
            elif 'סרניה' in pl: expanded.append('זריקת סרניה')
            elif 'פמוטידין' in pl: expanded.append('זריקת פמוטידין')
            elif 'סטרואיד' in pl: expanded.append('זריקת סטרואידים')
            else: expanded.append(f'זריקת {p.strip()}')
        return expanded
    return None

# ============================================================
# FILE LOADING — auto-detect structure
# ============================================================
def load_clinic(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        for enc in ['utf-16','utf-8-sig','utf-8','cp1255']:
            for sep in ['\t',',']:
                try:
                    df = pd.read_csv(path, encoding=enc, sep=sep, nrows=3)
                    if len(df.columns) > 2:
                        df = pd.read_csv(path, encoding=enc, sep=sep)
                        break
                except: continue
            else: continue
            break
    else:
        df = pd.read_excel(path)

    # Map columns — order matters: check specific before generic
    col_map = {}
    for c in df.columns:
        cl = str(c).strip()
        if cl == 'שם הלקוח' or cl == 'לקוח': col_map['client'] = c
        elif cl == 'שם החיה' or cl == 'חיה': col_map['animal'] = c
        elif 'תאריך' in cl and 'תאריך' not in col_map: col_map['date'] = c
        elif cl == 'פריט' or cl == 'טיפול': col_map['treatment'] = c
        elif cl == 'כמות': col_map['qty'] = c
    # Price: look for מחיר but NOT if already mapped as treatment
    for c in df.columns:
        cl = str(c).strip()
        if 'מחיר' in cl: col_map['price'] = c

    df['client'] = df[col_map['client']].apply(clean)
    df['animal'] = df[col_map.get('animal', col_map['client'])].apply(clean)
    df['date'] = pd.to_datetime(df[col_map['date']].apply(clean), format='%d/%m/%Y', errors='coerce')
    if df['date'].isna().all():
        df['date'] = pd.to_datetime(df[col_map['date']], errors='coerce')
    df['treatment'] = df[col_map['treatment']].apply(clean)
    df['price'] = pd.to_numeric(df[col_map['price']], errors='coerce').fillna(0) if 'price' in col_map else 0

    df = df.dropna(subset=['date'])
    if df.empty:
        return pd.DataFrame(columns=['client','animal','date','treatment','norm','price'])
    df = df[df['treatment'].str.strip() != ''].copy()
    if df.empty:
        return pd.DataFrame(columns=['client','animal','date','treatment','norm','price'])
    df = df[~df['treatment'].apply(is_excluded)].copy()
    if df.empty:
        return pd.DataFrame(columns=['client','animal','date','treatment','norm','price'])
    # Also exclude vaccine items (handled by vaccines audit skill)
    vaccine_patterns = ['תולעת הפארק','תילוע','כלבת','משושה','מרובע','מתומן','dp plus','dp +']
    df = df[~df['treatment'].str.lower().apply(lambda x: any(v in x for v in vaccine_patterns))].copy()
    if df.empty:
        return pd.DataFrame(columns=['client','animal','date','treatment','norm','price'])
    df['norm'] = df['treatment'].apply(normalize_treatment)
    return df[['client','animal','date','treatment','norm','price']].reset_index(drop=True)

def load_insurer(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        for enc in ['utf-16','utf-8-sig','utf-8','cp1255']:
            for sep in ['\t',',']:
                try:
                    df = pd.read_csv(path, encoding=enc, sep=sep, nrows=3)
                    if len(df.columns) > 2:
                        df = pd.read_csv(path, encoding=enc, sep=sep)
                        break
                except: continue
            else: continue
            break
    else:
        df = pd.read_excel(path)

    # Detect columns
    has_split_name = any('שם פרטי' in str(c) for c in df.columns)
    col_map = {}
    for c in df.columns:
        cl = str(c).strip()
        if 'שם פרטי' in cl: col_map['first'] = c
        elif 'שם משפחה' in cl or 'משפחה' in cl: col_map['last'] = c
        elif cl in ('שם הלקוח', 'לקוח', 'שם לקוח'): col_map['client'] = c
        elif cl in ('שם החיה', 'חיה', 'שם בעח', 'שם בע"ח'): col_map['animal'] = c
        elif 'תאריך' in cl: col_map['date'] = c
        elif 'רשימת טיפולים' in cl or 'טיפול' in cl or 'פריט' in cl: col_map['treatments'] = c

    if has_split_name:
        df['client'] = (df[col_map['first']].astype(str).str.strip() + ' ' + df[col_map['last']].astype(str).str.strip()).apply(clean)
    else:
        df['client'] = df[col_map['client']].apply(clean)

    df['date'] = pd.to_datetime(df[col_map['date']], errors='coerce')

    # Expand treatments (comma-separated) and injection expansions
    rows = []
    for _, row in df.iterrows():
        raw = str(row[col_map['treatments']])
        if raw in ['nan','NaT','']: continue
        animal = clean(str(row[col_map['animal']])) if 'animal' in col_map else ''
        for t in raw.split(','):
            t = clean(t)
            if not t: continue
            # Try expanding "זריקות (X, Y)"
            expanded = expand_insurer_treatment(t)
            if expanded:
                for e in expanded:
                    rows.append({'client': row['client'], 'animal': animal, 'date': row['date'],
                                 'treatment': t, 'norm': normalize_treatment(e), 'expanded': e})
            else:
                if is_excluded(t): continue
                rows.append({'client': row['client'], 'animal': animal, 'date': row['date'],
                             'treatment': t, 'norm': normalize_treatment(t), 'expanded': t})

    if not rows:
        return pd.DataFrame(columns=['client','animal','date','treatment','norm','expanded'])
    return pd.DataFrame(rows).reset_index(drop=True)

def _suspect_reason(c_client, i_client, c_norm, i_norm, c_date, i_date, c_animal='', i_animal=''):
    """Generate human-readable reason for suspected match."""
    reasons = []
    ns = sim(c_client, i_client)
    if ns >= 0.80:
        reasons.append(f'שם לקוח דומה ({ns:.0%})')
    elif ns >= 0.55:
        reasons.append(f'שם משפחה דומה ({ns:.0%})')
    if c_animal and i_animal:
        an_sim = sim(c_animal, i_animal)
        if an_sim >= 0.80:
            reasons.append(f'שם חיה דומה ({an_sim:.0%})')
    date_diff = abs((c_date - i_date).days)
    if date_diff == 0:
        reasons.append('תאריך זהה')
    elif date_diff <= 2:
        reasons.append(f'תאריך קרוב (±{date_diff})')
    if c_norm == i_norm:
        reasons.append('טיפול זהה')
    elif sim(c_norm, i_norm) >= 0.6:
        reasons.append('טיפול דומה')
    elif ('זריקת' in c_norm and 'זריקת' in i_norm) or ('זריקת' in c_norm and i_norm in ['סרניה','פמוטידין','כאבים','אנטיביוטיקה']):
        reasons.append('סוג זריקה תואם')
    elif is_bundle(c_norm) and ('דם' in i_norm or is_bundle(i_norm)):
        reasons.append('בדיקת דם תואמת')
    elif 'ספירה' in c_norm and ('דם' in i_norm or 'משטח' in i_norm):
        reasons.append('בדיקת דם תואמת')
    return ' | '.join(reasons) if reasons else 'חפיפה חלקית'

# ============================================================
# MATCHING ENGINE
# ============================================================
def run_audit(clinic, insurer):
    if clinic.empty and insurer.empty:
        empty_t1 = pd.DataFrame(columns=['שם לקוח','שם חיה','תאריך','טיפול במרפאה','מחיר','טיפול מנורמל','סטטוס'])
        empty_t3 = pd.DataFrame()
        empty_t4 = pd.DataFrame()
        empty_t5 = pd.DataFrame()
        clinic['matched'] = pd.Series(dtype=bool)
        clinic['match_type'] = pd.Series(dtype=str)
        clinic['match_note'] = pd.Series(dtype=str)
        insurer['matched'] = pd.Series(dtype=bool)
        return empty_t1, empty_t3, empty_t4, empty_t5, clinic, insurer
    clinic['matched'] = False
    clinic['match_type'] = ''
    clinic['match_note'] = ''
    insurer['matched'] = False

    delta_rows = []  # fuzzy/bundle matches
    partial_rows = []  # partial matches

    # --- Pass 1A: SPECIFIC matching (exact/fuzzy, non-generic items) ---
    for ci in clinic.index:
        if clinic.at[ci, 'matched']: continue
        cn, cd, ct = clinic.at[ci, 'client'], clinic.at[ci, 'date'], clinic.at[ci, 'norm']

        for ii in insurer.index:
            if insurer.at[ii, 'matched']: continue
            if insurer.at[ii, 'norm'] in GENERIC_NORMS: continue  # skip generic for now
            if abs((cd - insurer.at[ii, 'date']).days) > TOLERANCE_DAYS: continue
            if sim(cn, insurer.at[ii, 'client']) < FUZZY: continue
            if not animal_compatible(clinic.at[ci,'animal'], insurer.at[ii,'animal']): continue

            if treatments_match(ct, insurer.at[ii, 'norm']):
                mtype = 'exact' if sim(cn, insurer.at[ii,'client']) == 1.0 and ct == insurer.at[ii,'norm'] else 'fuzzy'
                clinic.at[ci, 'matched'] = True
                clinic.at[ci, 'match_type'] = mtype
                insurer.at[ii, 'matched'] = True
                if mtype == 'fuzzy':
                    delta_rows.append({
                        'לקוח_מרפאה': cn, 'חיה_מרפאה': clinic.at[ci,'animal'],
                        'לקוח_מרפאט': insurer.at[ii,'client'],
                        'תאריך': cd.strftime('%d/%m/%Y'),
                        'טיפול_מרפאה': clinic.at[ci,'treatment'],
                        'טיפול_מרפאט': insurer.at[ii,'treatment']})
                break

    # --- Pass 1B: GENERIC matching (זריקות, בדיקה, etc.) — 1:1 ---
    for ci in clinic.index:
        if clinic.at[ci, 'matched']: continue
        cn, cd, ct = clinic.at[ci, 'client'], clinic.at[ci, 'date'], clinic.at[ci, 'norm']

        for ii in insurer.index:
            if insurer.at[ii, 'matched']: continue
            if insurer.at[ii, 'norm'] not in GENERIC_NORMS: continue  # only generic now
            if abs((cd - insurer.at[ii, 'date']).days) > TOLERANCE_DAYS: continue
            if sim(cn, insurer.at[ii, 'client']) < FUZZY: continue
            if not animal_compatible(clinic.at[ci,'animal'], insurer.at[ii,'animal']): continue

            if treatments_match(ct, insurer.at[ii, 'norm']):
                mtype = 'fuzzy'
                clinic.at[ci, 'matched'] = True
                clinic.at[ci, 'match_type'] = mtype
                insurer.at[ii, 'matched'] = True
                delta_rows.append({
                    'לקוח_מרפאה': cn, 'חיה_מרפאה': clinic.at[ci,'animal'],
                    'לקוח_מרפאט': insurer.at[ii,'client'],
                    'תאריך': cd.strftime('%d/%m/%Y'),
                    'טיפול_מרפאה': clinic.at[ci,'treatment'],
                    'טיפול_מרפאט': insurer.at[ii,'treatment']})
                break

    # --- Pass 2: Bundle matching ---
    for ci in clinic.index:
        if clinic.at[ci, 'matched']: continue
        cn, cd, ct = clinic.at[ci, 'client'], clinic.at[ci, 'date'], clinic.at[ci, 'norm']

        for ii in insurer.index:
            # Don't require insurer unmatched for bundles (one bundle covers multiple clinic rows)
            if abs((cd - insurer.at[ii, 'date']).days) > TOLERANCE_DAYS: continue
            if sim(cn, insurer.at[ii, 'client']) < FUZZY: continue
            if not animal_compatible(clinic.at[ci,'animal'], insurer.at[ii,'animal']): continue
            i_norm = insurer.at[ii, 'norm']

            # Case A: insurer has bundle, clinic has component
            if is_bundle(i_norm) and bundle_contains(i_norm, ct):
                clinic.at[ci, 'matched'] = True
                clinic.at[ci, 'match_type'] = 'bundle'
                clinic.at[ci, 'match_note'] = f'חלק מחבילה: {insurer.at[ii,"treatment"]}'
                delta_rows.append({
                    'לקוח_מרפאה': cn, 'חיה_מרפאה': clinic.at[ci,'animal'],
                    'לקוח_מרפאט': insurer.at[ii,'client'],
                    'תאריך': cd.strftime('%d/%m/%Y'),
                    'טיפול_מרפאה': clinic.at[ci,'treatment'],
                    'טיפול_מרפאט': insurer.at[ii,'treatment']})
                break

            # Case B: clinic has bundle, insurer has component
            if is_bundle(ct) and bundle_contains(ct, i_norm):
                clinic.at[ci, 'matched'] = True
                clinic.at[ci, 'match_type'] = 'bundle'
                insurer.at[ii, 'matched'] = True
                delta_rows.append({
                    'לקוח_מרפאה': cn, 'חיה_מרפאה': clinic.at[ci,'animal'],
                    'לקוח_מרפאט': insurer.at[ii,'client'],
                    'תאריך': cd.strftime('%d/%m/%Y'),
                    'טיפול_מרפאה': clinic.at[ci,'treatment'],
                    'טיפול_מרפאט': insurer.at[ii,'treatment']})
                break

    # --- Pass 3: Owner Switch (different family member, same last name + date ±1) ---
    for ci in clinic.index:
        if clinic.at[ci, 'matched']: continue
        cn, cd, ct = clinic.at[ci, 'client'], clinic.at[ci, 'date'], clinic.at[ci, 'norm']
        c_last = cn.split()[-1] if cn.split() else cn

        for ii in insurer.index:
            if insurer.at[ii, 'matched']: continue
            if abs((cd - insurer.at[ii, 'date']).days) > 1: continue
            in_ = insurer.at[ii, 'client']
            i_last = in_.split()[-1] if in_.split() else in_
            last_sim = sim(c_last, i_last)
            full_sim = sim(cn, in_)
            # Same last name (≥80%) but different full name (below FUZZY threshold)
            if last_sim < 0.80 or full_sim >= FUZZY: continue
            if not animal_compatible(clinic.at[ci,'animal'], insurer.at[ii,'animal']): continue

            if treatments_match(ct, insurer.at[ii, 'norm']):
                clinic.at[ci, 'matched'] = True
                clinic.at[ci, 'match_type'] = 'owner_switch'
                clinic.at[ci, 'match_note'] = f'החלפת בעלים: {in_}'
                insurer.at[ii, 'matched'] = True
                delta_rows.append({
                    'לקוח_מרפאה': cn, 'חיה_מרפאה': clinic.at[ci, 'animal'],
                    'לקוח_מרפאט': in_,
                    'תאריך': cd.strftime('%d/%m/%Y'),
                    'טיפול_מרפאה': clinic.at[ci, 'treatment'],
                    'טיפול_מרפאט': insurer.at[ii, 'treatment']})
                break

    # --- Pass 4: Animal name matching (catches mismatched client names, same pet) ---
    for ci in clinic.index:
        if clinic.at[ci, 'matched']: continue
        cn, cd, ct = clinic.at[ci, 'client'], clinic.at[ci, 'date'], clinic.at[ci, 'norm']
        c_animal = clinic.at[ci, 'animal']
        if not c_animal: continue

        for ii in insurer.index:
            if insurer.at[ii, 'matched']: continue
            if abs((cd - insurer.at[ii, 'date']).days) > REVERSE_TOLERANCE: continue
            i_animal = insurer.at[ii, 'animal']
            if not i_animal: continue
            if sim(c_animal, i_animal) < FUZZY: continue

            if treatments_match(ct, insurer.at[ii, 'norm']):
                clinic.at[ci, 'matched'] = True
                clinic.at[ci, 'match_type'] = 'animal_match'
                clinic.at[ci, 'match_note'] = f'התאמת שם חיה: {i_animal} ← {insurer.at[ii, "client"]}'
                insurer.at[ii, 'matched'] = True
                delta_rows.append({
                    'לקוח_מרפאה': cn, 'חיה_מרפאה': c_animal,
                    'לקוח_מרפאט': insurer.at[ii, 'client'],
                    'תאריך': cd.strftime('%d/%m/%Y'),
                    'טיפול_מרפאה': clinic.at[ci, 'treatment'],
                    'טיפול_מרפאט': insurer.at[ii, 'treatment']})
                break

    # --- Build output tables ---
    # Table 5: Suspected matches — unmatched clinic items with partial evidence in insurer
    # Criteria: same date ±2 days + client sim >= 0.55 + (treatment sim >= 0.5 OR same norm category)
    suspect_rows = []
    for ci in clinic.index:
        if clinic.at[ci, 'matched']: continue
        cn, cd, ct = clinic.at[ci, 'client'], clinic.at[ci, 'date'], clinic.at[ci, 'norm']
        best_candidate = None
        best_score = 0
        c_animal = clinic.at[ci, 'animal']
        for ii in insurer.index:
            date_diff = abs((cd - insurer.at[ii, 'date']).days)
            if date_diff > TOLERANCE_DAYS: continue
            name_sim = sim(cn, insurer.at[ii, 'client'])
            i_animal = insurer.at[ii, 'animal']
            # Require at least one of: name similarity OR animal name similarity
            animal_sim = sim(c_animal, i_animal) if c_animal and i_animal else 0
            if name_sim < 0.55 and animal_sim < FUZZY: continue
            i_norm = insurer.at[ii, 'norm']
            # Score: name similarity + treatment similarity + date closeness + animal bonus
            treat_sim = sim(ct, i_norm)
            # Bonus for same category (both injections, both blood tests, etc.)
            category_bonus = 0
            if 'זריקת' in ct and 'זריקת' in i_norm: category_bonus = 0.3
            if 'זריקת' in ct and i_norm in ['זריקה_כללית','סרניה','פמוטידין','כאבים','אנטיביוטיקה']: category_bonus = 0.3
            if is_bundle(ct) and (is_bundle(i_norm) or 'דם' in i_norm): category_bonus = 0.3
            if 'ספירה' in ct and ('משטח' in i_norm or 'דם' in i_norm or 'ספירה' in i_norm): category_bonus = 0.3
            if ct == 'ביקור' and i_norm in ['בדיקה','ביקור']: category_bonus = 0.3
            if 'מומחה' in ct and 'מומחה' in i_norm: category_bonus = 0.3
            animal_bonus = 0.3 if animal_sim >= FUZZY else (0.15 if animal_sim >= 0.60 else 0)
            score = name_sim * 0.4 + treat_sim * 0.3 + category_bonus + animal_bonus + (1 - date_diff / 7) * 0.1
            if score > best_score and score >= 0.45:
                best_score = score
                best_candidate = ii

        if best_candidate is not None:
            ii = best_candidate
            suspect_rows.append({
                'לקוח_מרפאה': cn, 'חיה_מרפאה': c_animal,
                'לקוח_מרפאט': insurer.at[ii, 'client'],
                'תאריך_מרפאה': cd.strftime('%d/%m/%Y'),
                'תאריך_מרפאט': insurer.at[ii, 'date'].strftime('%d/%m/%Y'),
                'טיפול_מרפאה': clinic.at[ci, 'treatment'],
                'טיפול_מרפאט': insurer.at[ii, 'treatment'],
                'סיבת_חשד': _suspect_reason(cn, insurer.at[ii,'client'], ct, insurer.at[ii,'norm'], cd, insurer.at[ii,'date'], c_animal, insurer.at[ii,'animal'])
            })
    table5 = pd.DataFrame(suspect_rows)

    # Table 1: Lost claims (remove suspected matches — they go to table 5)
    lost = clinic[~clinic['matched']].copy()
    table1 = lost[['client','animal','date','treatment','price','norm']].copy()
    table1.columns = ['שם לקוח','שם חיה','תאריך','טיפול במרפאה','מחיר','טיפול מנורמל']
    if not table1.empty and pd.api.types.is_datetime64_any_dtype(table1['תאריך']):
        table1['תאריך'] = table1['תאריך'].dt.strftime('%d/%m/%Y')
    table1['סטטוס'] = 'לא נמצא בביטוח'

    # Table 3: Delta
    table3 = pd.DataFrame(delta_rows)

    # Table 4: Reverse orphans (insurer without clinic match)
    # Exclude vaccine items — those belong to the vaccines audit skill
    VACCINE_NORMS = ['תולעת הפארק','תילוע','חיסון כלבת','חיסון משושה','כלבת','משושה','מרובע','מתומן']
    table4_rows = []
    for ii in insurer.index:
        if insurer.at[ii, 'matched']: continue
        if any(v in insurer.at[ii, 'norm'] for v in VACCINE_NORMS): continue
        if any(v in insurer.at[ii, 'treatment'] for v in ['תולעת','תילוע','כלבת','משושה','מרובע','מתומן']): continue
        found = False
        for ci in clinic.index:
            if abs((insurer.at[ii,'date'] - clinic.at[ci,'date']).days) > REVERSE_TOLERANCE: continue
            if sim(insurer.at[ii,'client'], clinic.at[ci,'client']) >= FUZZY:
                if treatments_match(insurer.at[ii,'norm'], clinic.at[ci,'norm']):
                    found = True; break
        if not found:
            table4_rows.append({
                'שם לקוח (מרפאט)': insurer.at[ii,'client'],
                'שם חיה (מרפאט)': insurer.at[ii,'animal'],
                'תאריך': insurer.at[ii,'date'].strftime('%d/%m/%Y'),
                'טיפול (מרפאט)': insurer.at[ii,'treatment'],
                'טיפול מנורמל': insurer.at[ii,'norm'],
                'הערה': 'קיים במרפאט ללא התאמה במרפאה'})
    table4 = pd.DataFrame(table4_rows)

    # --- Cross-check: table1 (lost claims) ↔ table4 (orphans) by animal name ---
    # Same animal + same date in both = likely same visit with mismatched client names
    # Move such pairs to table5 (suspects) to avoid false double-listing
    if len(table1) > 0 and len(table4) > 0 and 'שם חיה (מרפאט)' in (table4.columns if len(table4) > 0 else []):
        t1_drop, t4_drop, cross_rows = set(), set(), []
        for i1, r1 in table1.iterrows():
            a1 = str(r1.get('שם חיה', '')).strip()
            if not a1: continue
            try: d1 = pd.to_datetime(r1['תאריך'], dayfirst=True)
            except: continue
            for i4, r4 in table4.iterrows():
                a4 = str(r4.get('שם חיה (מרפאט)', '')).strip()
                if not a4: continue
                if sim(a1, a4) < FUZZY: continue
                try: d4 = pd.to_datetime(r4['תאריך'], dayfirst=True)
                except: continue
                if abs((d1 - d4).days) > 2: continue
                t1_drop.add(i1)
                t4_drop.add(i4)
                cross_rows.append({
                    'לקוח_מרפאה': r1['שם לקוח'],
                    'חיה_מרפאה': a1,
                    'לקוח_מרפאט': r4['שם לקוח (מרפאט)'],
                    'תאריך_מרפאה': r1['תאריך'],
                    'תאריך_מרפאט': r4['תאריך'],
                    'טיפול_מרפאה': r1['טיפול במרפאה'],
                    'טיפול_מרפאט': r4['טיפול (מרפאט)'],
                    'סיבת_חשד': f'אותה חיה ({a1}/{a4}) — אבודה+יתום באותו תאריך'
                })
        if cross_rows:
            cross5 = pd.DataFrame(cross_rows)
            cross5.columns = ['לקוח_מרפאה','חיה_מרפאה','לקוח_מרפאט','תאריך_מרפאה','תאריך_מרפאט','טיפול_מרפאה','טיפול_מרפאט','סיבת_חשד']
            table1 = table1.drop(index=list(t1_drop)).reset_index(drop=True)
            table4 = table4.drop(index=list(t4_drop)).reset_index(drop=True)
            table5 = pd.concat([table5, cross5], ignore_index=True) if len(table5) > 0 else cross5

    return table1, table3, table4, table5, clinic, insurer

# ============================================================
# EXCEL OUTPUT
# ============================================================
def write_excel(table1, table3, table4, table5, clinic, insurer, output_path):
    wb = Workbook()
    hf = Font(bold=True, size=11, color='FFFFFF', name='Arial')
    df_ = Font(size=10, name='Arial')
    bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    ca = Alignment(horizontal='center', vertical='center')
    ra = Alignment(horizontal='right', vertical='center')

    def sheet(ws, title, tc, hc, headers, data, rf):
        ws.sheet_view.rightToLeft = True
        nc = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
        ws['A1'].value = title
        ws['A1'].font = Font(bold=True, size=14, color=tc, name='Arial')
        ws['A1'].alignment = ca
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=i, value=h)
            c.font = hf; c.fill = PatternFill('solid', fgColor=hc); c.alignment = ca; c.border = bd
        for ri, row in enumerate(data, 3):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = df_; c.alignment = ra; c.border = bd; c.fill = PatternFill('solid', fgColor=rf)
        for i in range(1, nc + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

    # Sheet 1: Lost claims
    ws1 = wb.active; ws1.title = 'תביעות אבודות'
    h1 = ['שם לקוח','שם חיה','תאריך','טיפול במרפאה','מחיר','סטטוס']
    d1 = [[r['שם לקוח'], r['שם חיה'], r['תאריך'], r['טיפול במרפאה'], r['מחיר'], r['סטטוס']] for _, r in table1.iterrows()]
    sheet(ws1, f'תביעות אבודות — {len(table1)} טיפולים חסרים בביטוח', 'CC0000', 'C00000', h1, d1, 'FFC7CE')

    # Sheet 2: Delta
    ws3 = wb.create_sheet('דוח דלתא')
    h3 = ['לקוח (מרפאה)','חיה (מרפאה)','לקוח (מרפאט)','תאריך','טיפול מרפאה','טיפול מרפאט']
    d3 = [list(r) for _, r in table3.iterrows()] if len(table3) > 0 else []
    sheet(ws3, f'דוח דלתא — התאמות פונטיות/חבילות ({len(table3)})', '006100', '006100', h3, d3, 'C6EFCE')

    # Sheet 3: Reverse orphans
    ws4 = wb.create_sheet('יתומים במרפאט')
    h4 = ['שם לקוח (מרפאט)','שם חיה (מרפאט)','תאריך','טיפול (מרפאט)','טיפול מנורמל','הערה']
    d4 = [[r.get('שם לקוח (מרפאט)',''), r.get('שם חיה (מרפאט)',''), r.get('תאריך',''), r.get('טיפול (מרפאט)',''), r.get('טיפול מנורמל',''), r.get('הערה','')] for _, r in table4.iterrows()] if len(table4) > 0 else []
    sheet(ws4, f'יתומים במרפאט — {len(table4)} רשומות', '7030A0', '7030A0', h4, d4, 'E8D5F5')

    # Sheet 4: Suspected matches
    ws6 = wb.create_sheet('חשד להתאמה')
    h6 = ['לקוח (מרפאה)','חיה (מרפאה)','לקוח (מרפאט)','תאריך מרפאה','תאריך מרפאט','טיפול מרפאה','טיפול מרפאט','סיבת חשד']
    d6 = [list(r) for _, r in table5.iterrows()] if len(table5) > 0 else []
    sheet(ws6, f'חשד להתאמה — {len(table5)} רשומות לבדיקה ידנית', 'E36C09', 'E36C09', h6, d6, 'FCE4D6')

    # Sheet 4: Summary
    ws5 = wb.create_sheet('סיכום'); ws5.sheet_view.rightToLeft = True
    ws5.column_dimensions['A'].width = 50; ws5.column_dimensions['B'].width = 15
    ws5.merge_cells('A1:B1')
    ws5['A1'].value = 'סיכום ביקורת טיפולים'
    ws5['A1'].font = Font(bold=True, size=14, name='Arial'); ws5['A1'].alignment = ca

    total = len(clinic)
    matched = clinic['matched'].sum()
    exact = len(clinic[clinic['match_type'] == 'exact'])
    fuzzy = len(clinic[clinic['match_type'] == 'fuzzy'])
    bundle = len(clinic[clinic['match_type'] == 'bundle'])
    owner_sw = len(clinic[clinic['match_type'] == 'owner_switch'])
    animal_m = len(clinic[clinic['match_type'] == 'animal_match'])
    lost_val = table1['מחיר'].sum() if len(table1) > 0 else 0

    summary = [
        ('סה"כ פריטי טיפול במרפאה (אחרי סינון)', total),
        ('התאמה מלאה', exact),
        ('התאמה פונטית / גמישה', fuzzy),
        ('התאמת חבילות (Bundles)', bundle),
        ('התאמת החלפת בעלים', owner_sw),
        ('התאמת שם חיה', animal_m),
        ('סה"כ הותאמו', int(matched)),
        ('', ''),
        ('תביעות אבודות (חסרים בביטוח)', len(table1)),
        ('מתוכן — חשד להתאמה (לבדיקה ידנית)', len(table5)),
        ('שווי תביעות אבודות (₪)', f'{lost_val:,.0f}'),
        ('יתומים במרפאט', len(table4)),
    ]
    for i, (label, val) in enumerate(summary, 3):
        c1 = ws5.cell(row=i, column=1, value=label)
        c2 = ws5.cell(row=i, column=2, value=val)
        bold = 'סה"כ' in str(label) or 'אבודות' in str(label) or 'יתומים' in str(label)
        c1.font = Font(size=11, name='Arial', bold=bold)
        c2.font = Font(size=11, name='Arial', bold=True); c2.alignment = ca
        if 'אבודות' in str(label):
            c1.font = Font(size=11, name='Arial', bold=True, color='CC0000')
            c2.font = Font(size=11, name='Arial', bold=True, color='CC0000')

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Total: {total} | Matched: {int(matched)} (exact:{exact} fuzzy:{fuzzy} bundle:{bundle} owner_switch:{owner_sw} animal:{animal_m}) | Lost: {len(table1)} (₪{lost_val:,.0f}) | Reverse orphans: {len(table4)}")

# ============================================================
if __name__ == '__main__':
    if len(sys.argv) < 4:
        print("Usage: treatments_audit.py <clinic_file> <insurer_file> <output.xlsx>")
        sys.exit(1)
    clinic = load_clinic(sys.argv[1])
    insurer = load_insurer(sys.argv[2])
    t1, t3, t4, t5, clinic, insurer = run_audit(clinic, insurer)
    write_excel(t1, t3, t4, t5, clinic, insurer, sys.argv[3])
